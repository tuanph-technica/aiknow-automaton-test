import os
import io
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
import logging
from typing import List, Dict

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ChatBotResultWriter:
    """
    Excel writer for ChatBot test results with image support

    This class creates comprehensive test result reports with:
    - Test data and results
    - Screenshots for evidence
    - Pass/Fail statistics
    - Hyperlinks to evidence sheet
    """

    def __init__(self, filename: str):
        """
        Initialize the result writer

        Args:
            filename: Path to the output Excel file
        """
        self.filename = filename
        self.wb = openpyxl.Workbook()

        # Create sheets
        self.ws_results = self.wb.active
        self.ws_results.title = "Test Results"

        self.ws_evidence = self.wb.create_sheet("Evidence")
        self.ws_summary = self.wb.create_sheet("Summary")

        self.current_row = 1
        self.evidence_row = 1

        logger.info(f"ChatBotResultWriter initialized: {filename}")

    def _style_header(self, cell):
        """Apply header styling to a cell"""
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_cell(self, cell, is_pass=None):
        """Apply cell styling"""
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if is_pass is not None:
            if is_pass:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def write_headers(self):
        """Write column headers to the results sheet"""
        headers = [
            "STT",
            "Test Data Describe",
            "Data Type",
            "Question",
            "Expected Context",
            "Expected Result",
            "Actual Response",
            "Actual Context",
            "Test Result",
            "Evaluation Score",
            "Evaluation Details",
            "Response Time",
            "Model",
            "Evidence"
        ]

        for col, header in enumerate(headers, 1):
            cell = self.ws_results.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Set column widths
        self.ws_results.column_dimensions['A'].width = 8
        self.ws_results.column_dimensions['B'].width = 25
        self.ws_results.column_dimensions['C'].width = 15
        self.ws_results.column_dimensions['D'].width = 40
        self.ws_results.column_dimensions['E'].width = 30
        self.ws_results.column_dimensions['F'].width = 30
        self.ws_results.column_dimensions['G'].width = 40
        self.ws_results.column_dimensions['H'].width = 30
        self.ws_results.column_dimensions['I'].width = 12
        self.ws_results.column_dimensions['J'].width = 12
        self.ws_results.column_dimensions['K'].width = 25
        self.ws_results.column_dimensions['L'].width = 15
        self.ws_results.column_dimensions['M'].width = 20
        self.ws_results.column_dimensions['N'].width = 15

        self.current_row = 2
        logger.info("Headers written to results sheet")

    def add_result_row(self, result_data: Dict):
        """
        Add a test result row

        Args:
            result_data: Dictionary containing test result data
        """
        row = self.current_row

        # Write data cells
        data_mapping = [
            ("STT", 1),
            ("Test Data Describe", 2),
            ("Data Type", 3),
            ("question", 4),
            ("expected_context", 5),
            ("expected_result", 6),
            ("actual_response", 7),
            ("actual_context", 8),
            ("test_result", 9),
            ("evaluation_score", 10),
            ("evaluation_details", 11),
            ("time_response", 12),
            ("model", 13)
        ]

        is_pass = result_data.get("test_result", "").lower() == "pass"

        for key, col in data_mapping:
            cell = self.ws_results.cell(row=row, column=col)
            value = result_data.get(key, "")

            # Format score as percentage
            if key == "evaluation_score" and isinstance(value, (int, float)):
                cell.value = f"{value:.1%}"
            else:
                cell.value = value

            # Apply styling
            if col == 9:  # Test Result column
                self._style_cell(cell, is_pass)
            else:
                self._style_cell(cell)

        # Add evidence screenshot
        evidence_cell = self.ws_results.cell(row=row, column=14)
        screenshot = result_data.get("evident")

        if screenshot:
            # Add screenshot to evidence sheet
            evidence_ref = self.ws_evidence.cell(row=self.evidence_row, column=1)
            self._add_screenshot(evidence_ref, screenshot)

            # Add hyperlink
            evidence_cell.value = "View Evidence"
            evidence_cell.hyperlink = Hyperlink(
                ref=f"Evidence!{evidence_ref.coordinate}",
                display="View Evidence"
            )
            evidence_cell.font = Font(color="0000FF", underline="single")

            self.evidence_row += 1
        else:
            evidence_cell.value = "No evidence"
            self._style_cell(evidence_cell)

        # Adjust row height
        self.ws_results.row_dimensions[row].height = 60

        self.current_row += 1

    def _add_screenshot(self, cell, screenshot_bytes, size=(400, 300)):
        """
        Add screenshot to evidence sheet

        Args:
            cell: Target cell for screenshot
            screenshot_bytes: Screenshot image bytes
            size: Image size (width, height)
        """
        try:
            image_stream = io.BytesIO(screenshot_bytes)
            img = Image(image_stream)
            img.width, img.height = size

            self.ws_evidence.add_image(img, cell.coordinate)

            # Adjust row height
            self.ws_evidence.row_dimensions[cell.row].height = size[1] * 0.75

        except Exception as e:
            logger.error(f"Failed to add screenshot: {e}")

    def write_summary(self, results: List[Dict]):
        """
        Write summary statistics

        Args:
            results: List of test result dictionaries
        """
        # Calculate statistics
        total_tests = len(results)
        passed = sum(1 for r in results if r.get("test_result", "").lower() == "pass")
        failed = sum(1 for r in results if r.get("test_result", "").lower() == "fail")
        errors = sum(1 for r in results if r.get("test_result", "").lower() in ["error", "timeout"])

        pass_rate = (passed / total_tests * 100) if total_tests > 0 else 0

        # Calculate average score
        scores = [r.get("evaluation_score", 0) for r in results if isinstance(r.get("evaluation_score"), (int, float))]
        avg_score = sum(scores) / len(scores) if scores else 0

        # Calculate average response time
        times = []
        for r in results:
            time_str = r.get("time_response", "")
            if "second" in time_str:
                try:
                    times.append(float(time_str.split()[0]))
                except:
                    pass
        avg_time = sum(times) / len(times) if times else 0

        # Write summary
        summary_data = [
            ("Test Summary", ""),
            ("", ""),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Tests", total_tests),
            ("Passed", passed),
            ("Failed", failed),
            ("Errors/Timeouts", errors),
            ("Pass Rate", f"{pass_rate:.1f}%"),
            ("", ""),
            ("Average Evaluation Score", f"{avg_score:.1%}"),
            ("Average Response Time", f"{avg_time:.2f} seconds"),
        ]

        for row_idx, (label, value) in enumerate(summary_data, 1):
            cell_label = self.ws_summary.cell(row=row_idx, column=1, value=label)
            cell_value = self.ws_summary.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:
                cell_label.font = Font(bold=True, size=14)
            elif label:
                cell_label.font = Font(bold=True)

        # Set column widths
        self.ws_summary.column_dimensions['A'].width = 30
        self.ws_summary.column_dimensions['B'].width = 20

        logger.info("Summary written successfully")

    def save(self):
        """Save the workbook"""
        try:
            # Freeze header row in results sheet
            self.ws_results.freeze_panes = "A2"

            # Save file
            self.wb.save(self.filename)
            logger.info(f"Results saved successfully: {self.filename}")

        except Exception as e:
            logger.error(f"Error saving results: {e}")
            raise


def export_chatbot_results(results: List[Dict], filename: str):
    """
    Export chatbot test results to Excel

    Args:
        results: List of test result dictionaries
        filename: Output Excel file path
    """
    if not results:
        logger.warning("No results to export")
        return

    try:
        writer = ChatBotResultWriter(filename)
        writer.write_headers()

        for result in results:
            writer.add_result_row(result)

        writer.write_summary(results)
        writer.save()

        logger.info(f"Successfully exported {len(results)} results to {filename}")

    except Exception as e:
        logger.error(f"Failed to export results: {e}")
        raise


if __name__ == "__main__":
    # Example usage
    sample_results = [
        {
            "STT": 1,
            "Test Data Describe": "Test case 1",
            "Data Type": "Question",
            "question": "What is Python?",
            "expected_context": "Programming language",
            "expected_result": "Python is a programming language",
            "actual_response": "Python is a high-level programming language",
            "actual_context": "Programming context",
            "test_result": "pass",
            "evaluation_score": 0.85,
            "evaluation_details": "Good match",
            "time_response": "2.5 seconds",
            "model": "GPT-4",
            "evident": None
        }
    ]

    export_chatbot_results(sample_results, "test_output.xlsx")
