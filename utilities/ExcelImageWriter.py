import io

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.hyperlink import Hyperlink
import os
from pathlib import Path
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from base.base_driver import BaseDriver


class ExcelImageWriter(BaseDriver):
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "TestResult"
        self.sheet_evident = self.wb.create_sheet("Evidents")
        self.current_row = 1

        # Set up logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    def add_hyper_link(self,current_cell,reference_cell):
        sheet_name = reference_cell.parent.title
        cell_coordinate = reference_cell.coordinate
        full_address = f"{sheet_name}!{cell_coordinate}"
        current_cell.hyperlink = Hyperlink(ref=full_address, display="see evident")
        current_cell.font = Font(color="0000FF", underline="single")




    def write_headers(self, headers):
        """Write column headers with styling"""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)

            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        self.current_row = 2

    def add_row_with_image(self, row_data, image_column=None, image_size=(100, 100)):
        """Add a row of data with optional image"""
        for col, (key, value) in enumerate(row_data.items(), 1):
            cell = self.ws.cell(row=self.current_row, column=col)
            if key == image_column and value:
                cell_ref = self.sheet_evident.cell(row=self.current_row,column=1)
                success = self._add_image_to_cell(cell_ref, value, image_size)
                if success:
                    cell.value = "See evident"
                    self.add_hyper_link(current_cell=cell,reference_cell=cell_ref)

                else:
                    cell.value = "❌ Image Error"
            else:
                cell.value = value

        # Adjust row height for images
        if image_column and image_column in row_data:
            self.ws.row_dimensions[self.current_row].height = max(image_size[1] * 0.75, 20)

        self.current_row += 1

    def _add_image_to_cell(self, cell, image_bytes, size):
        """Add image to specific cell"""
        try:
            image_stream = io.BytesIO(image_bytes)
            img = Image(image_stream)
            img.width, img.height = size
            self.sheet_evident.add_image(img, cell.coordinate)
            return True
        except Exception as e:
            print({e})
            return False

    def save(self):
        """Save the workbook"""
        try:
            # Auto-adjust column widths
            for column in self.ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 30)
                self.ws.column_dimensions[column_letter].width = adjusted_width

            self.wb.save(self.filename)
            self.logger.info(f"Excel file saved successfully: {self.filename}")

        except Exception as e:
            self.logger.error(f"Error saving file: {str(e)}")


def export_data_to_excel(data_list, filename, image_column='image'):
    """Main function to export data with images"""

    if not data_list:
        print("No data to export")
        return

    writer = ExcelImageWriter(filename)

    # Write headers
    headers = list(data_list[0].keys())
    writer.write_headers(headers)

    # Write data rows
    for item in data_list:
        writer.add_row_with_image(item, image_column, image_size=(80, 80))

    # Save file
    writer.save()


