from pathlib import Path
from openpyxl.workbook import Workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from PIL import Image as PILImage
import io
import os



def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row
def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column
def readData(file,sheet_name,row_num,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num,column=columnno).value
def writeData(file,sheetName,row_num,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=row_num,column=columnno).value = data
    workbook.save(file)
def create_excel_test_file(data_list, filename, image_column='image'):
    wb = Workbook()
    ws = wb.active
    if not data_list:
        return
    # Get headers from first dictionary
    headers = list(data_list[0].keys())
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row in data_list:
        print(row)
    start_row = 2

    # Write data and handle images
    for row_idx, item in enumerate(data_list, 2):  # Start from row 2
        for col_idx, (key, value) in enumerate(item.items(), 1):
            if key == image_column and value:
                # Handle image field
                try:
                    if os.path.exists(str(value)):
                        # Resize image to fit cell
                        img = Image(value)
                        img.width = 100  # Adjust size as needed
                        img.height = 100

                        # Add image to cell
                        cell = ws.cell(row=row_idx, column=col_idx)
                        ws.add_image(img, cell.coordinate)

                        # Adjust row height to fit image
                        ws.row_dimensions[row_idx].height = 80

                        # Set cell value to image filename
                        cell.value = os.path.basename(str(value))
                    else:
                        ws.cell(row=row_idx, column=col_idx, value="Image not found")
                except Exception as e:
                    ws.cell(row=row_idx, column=col_idx, value=f"Error: {str(e)}")
            else:
                # Handle regular data
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(filename)
    print(f"Excel file saved: {filename}")






